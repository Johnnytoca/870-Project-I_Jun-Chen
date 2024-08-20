import time
import os
import paramiko
import openpyxl


def confirm_commands(cmd):
    """
    Double check commands are right
    :parameters:
        cmd(type list): Configuration commands
    :return:int: 1 for yes, 0 for no, None for the input error
    """
    print(f"Command: {cmd}")

    try:
        operate = int(input("Confirm to execute (1:yes 0:no):"))
    except ValueError:
        print("Input error")
        return None

    if operate not in [0, 1]:
        print("Input error")
        return None
    return operate


def connect_device(device_information, cmd):
    """
    Connect and config devices, then store log files in current catalogue
    :parameters:
        device_information(type dictionary): Including device_name:device_ip
        cmd(type list): Configuration commands
    """
    filedict = {}
    ssh_port = 22
    ssh_user = "huawei"
    ssh_passwd = "huawei"
    do_commands = cmd

    for device_name, device_ip in device_information.items():
        timestamp = time.strftime("%Y%m%d_%H%M%S")
        # Compatible with both Windows and Linux
        folder_path = os.path.join(os.getcwd(), f"{device_name}_{device_ip}_" + timestamp)
        os.makedirs(folder_path, exist_ok=True)
        filepath = os.path.join(folder_path, f"{device_name}_{device_ip}.txt")

        try:
            with open(filepath, "wb") as devicefile:
                filedict[device_name] = devicefile

                # Create a new session
                ssh_client = paramiko.SSHClient()
                # Specific RSA key policy: auto add the hash file for the public key of the SSH server
                ssh_client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
                # Specific SSH parameters
                ssh_client.connect(hostname=device_ip, port=ssh_port, username=ssh_user, password=ssh_passwd,
                                   look_for_keys=False)
                # Create a SSH terminal
                terminal = ssh_client.invoke_shell()
                # Send commands
                for each_command in do_commands:
                    terminal.send(each_command + "\n")
                    time.sleep(1)
                    standard_out = terminal.recv(65535)
                    devicefile.write(standard_out)

                print(f"{device_name} Done, close the session\n")
                terminal.close()
                ssh_client.close()
        except (paramiko.SSHException, IOError) as e:
            print(f"Can't connect to the device {device_name}({device_ip}): {str(e)}")

    # close all log files
    for fileclose in filedict.values():
        fileclose.close()


def dataload(file):
    """
    Step 1: Load data from the Excel file
    Step 2: Double-check if it is performed by calling confirm_commands function
    Step 3: Connect to devices and save the result to local files by calling connect_device function
    :parameter:
        file(type string): The Excel file path
    """
    workbook = openpyxl.load_workbook(excel_file)
    sheet_names = workbook.sheetnames
    print("Sheets' name:", sheet_names)

    sheet_connect = workbook[sheet_names[0]]
    sheet_cmd = workbook[sheet_names[1]]

    device_info = {}
    commands_list = []
    # Load commands from the excel
    for row in sheet_cmd.iter_rows(min_row=2, min_col=2, max_col=2):
        command = row[0].value  # 2nd col
        if command:
            commands_list.append(command)

    operation_tag = confirm_commands(commands_list)
    if operation_tag == 1:
        # Load SSH info from the excel
        for row in sheet_connect.iter_rows(min_row=2, min_col=2, max_col=3):
            device_name = row[0].value  # 2nd col
            device_ip = row[1].value  # 3rd col
            if device_name and device_ip:
                device_info[device_name] = device_ip
        connect_device(device_info, commands_list)
    else:
        print("Break")


if __name__ == '__main__':
    excel_file = input(r"Input the file path (e.g. /path/XXX.xlsx or X:\XXX.xlsx):")
    dataload(excel_file)
